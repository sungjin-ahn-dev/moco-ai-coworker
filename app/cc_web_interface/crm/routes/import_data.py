"""
데이터 임포트 API 라우트
Excel/JSON 파일에서 병원, HCP, 처방, 매출 등 벌크 임포트
"""

import logging
from typing import Optional

from fastapi import APIRouter, Depends, UploadFile, File, Form
from sqlalchemy import select, func
from sqlalchemy.ext.asyncio import AsyncSession

from app.cc_web_interface.crm.database import get_db
from app.cc_web_interface.crm.models import (
    Company, Contact, Prescription, PatientCompliance,
    SalesTransaction, ProductListing, KOLPlan, HospitalContract,
)
from app.cc_web_interface.crm.schemas import (
    ImportResult, SuccessResponse, ErrorResponse,
)

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/import", tags=["데이터 임포트"])


@router.post("/reload-db", response_model=SuccessResponse)
async def reload_db():
    """DB 엔진 연결 풀 리셋 (외부에서 직접 넣은 데이터 반영)"""
    from app.cc_web_interface.crm.database import engine
    await engine.dispose()
    return SuccessResponse(data={"message": "DB 연결 리셋 완료"})


def _parse_excel(file_bytes: bytes, sheet_name: Optional[str] = None,
                  max_col: Optional[int] = None) -> list[dict]:
    """Excel 파일을 파싱하여 딕셔너리 리스트로 반환.
    max_col: 사용할 최대 컬럼 수 (중복 헤더 방지용)
    """
    import openpyxl
    from io import BytesIO

    wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)

    if sheet_name:
        if sheet_name not in wb.sheetnames:
            wb.close()
            raise ValueError(f"시트 '{sheet_name}'을 찾을 수 없습니다. 가능한 시트: {wb.sheetnames}")
        ws = wb[sheet_name]
    else:
        ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 2:
        return []

    # 헤더 행 찾기 - 알려진 헤더 키워드가 있는 첫 행을 우선
    HEADER_KEYWORDS = {'병원', '병원명', '이름', '성명', '연도', '제품', '처방', 'No', 'No.',
                       '고유키값', '담당자', '의료진', '교수', '번호', '처방 ID', '환자 ID',
                       '병원 LIST', '채널(거래처)', '공급가', 'Ownership'}
    header_idx = 0
    max_str_count = 0
    best_keyword_row = None
    for i, row in enumerate(rows[:10]):
        str_vals = [str(v).strip() for v in row if isinstance(v, str) and v.strip()]
        str_count = len(str_vals)
        if best_keyword_row is None and any(sv in HEADER_KEYWORDS for sv in str_vals):
            best_keyword_row = i
        if str_count > max_str_count:
            max_str_count = str_count
            header_idx = i
    if best_keyword_row is not None:
        header_idx = best_keyword_row

    headers = rows[header_idx]
    # max_col로 헤더 잘라내기 (중복 헤더 방지)
    if max_col:
        headers = headers[:max_col]

    # 깨끗한 헤더 (None 제거, 줄바꿈 제거, 중복 제거)
    clean_headers = []
    seen = set()
    for h in headers:
        if h is not None:
            name = str(h).strip().replace('\n', ' ')
            # 중복 헤더명은 _2, _3 등 붙이기
            if name in seen:
                orig = name
                idx = 2
                while name in seen:
                    name = f"{orig}_{idx}"
                    idx += 1
            seen.add(name)
            clean_headers.append(name)
        else:
            clean_headers.append(None)

    records = []
    for row in rows[header_idx + 1:]:
        if all(v is None for v in row):
            continue
        record = {}
        cols = row[:max_col] if max_col else row
        for j, val in enumerate(cols):
            if j < len(clean_headers) and clean_headers[j] is not None:
                record[clean_headers[j]] = val
        if record:
            records.append(record)

    return records


def _parse_excel_raw(file_bytes: bytes, sheet_name: Optional[str] = None) -> list[list]:
    """Excel 파일을 raw 행 리스트로 반환 (특수 레이아웃용)"""
    import openpyxl
    from io import BytesIO

    wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.active
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()
    return rows


def _str(v) -> Optional[str]:
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def _int(v) -> Optional[int]:
    if v is None:
        return None
    try:
        return int(float(v))
    except (ValueError, TypeError):
        return None


def _float(v) -> Optional[float]:
    if v is None:
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


def _bool_ish(v) -> bool:
    """'O', 'Y', True, 1 등을 True로 변환"""
    if v is None:
        return False
    if isinstance(v, bool):
        return v
    s = str(v).strip().upper()
    return s in ('O', 'Y', 'YES', 'TRUE', '1', '완료')


# ────────────────── Phase 1: 병원 임포트 ──────────────────


@router.post("/hospitals", response_model=SuccessResponse)
async def import_hospitals(
    file: UploadFile = File(...),
    sheet_name: Optional[str] = Form(None),
    db: AsyncSession = Depends(get_db),
):
    """병원 마스터 데이터 벌크 임포트 (upsert by hospital_code or name)"""
    try:
        file_bytes = await file.read()
        records = _parse_excel(file_bytes, sheet_name)
    except Exception as e:
        return ErrorResponse(message=f"파일 파싱 실패: {str(e)}")

    result = ImportResult(total=len(records))

    # 제품 관련 컬럼명 (MASTER DB_Hospital의 실제 컬럼)
    PRODUCT_COLUMNS = [
        '제품A(도입)', '제품A', '제품C', '제품D',
        '유통A', '유통B', 'CAT/CLT', 'CAT', 'ProductD',
        'Listing Status', '완료일',
    ]

    for rec in records:
        try:
            # 병원명 매핑 (다양한 헤더명 지원)
            name = _str(rec.get('병원명') or rec.get('병원') or rec.get('실시기관명')
                       or rec.get('요양기관명') or rec.get('병원/약국명') or rec.get('name')
                       or rec.get('요양기관명칭') or rec.get('병원 LIST'))
            if not name:
                result.skipped += 1
                continue

            # hospital_code 매핑
            hospital_code = _str(rec.get('고유키값') or rec.get('요양기관번호')
                                or rec.get('요양기호') or rec.get('hospital_code'))

            # 기존 병원 찾기 (hospital_code 우선, 없으면 이름으로)
            existing = None
            if hospital_code:
                q = await db.execute(
                    select(Company).where(Company.hospital_code == hospital_code)
                )
                existing = q.scalar_one_or_none()

            if not existing:
                q = await db.execute(
                    select(Company).where(Company.name == name)
                )
                existing = q.scalar_one_or_none()

            hospital_type = _str(rec.get('구분') or rec.get('병원구분') or rec.get('종별')
                                or rec.get('요양종별') or rec.get('병원/약국구분')
                                or rec.get('hospital_type') or rec.get('병원구분'))
            region_1 = _str(rec.get('지역1') or rec.get('시도명') or rec.get('region_1'))
            region_2 = _str(rec.get('지역2') or rec.get('region_2'))
            region_3 = _str(rec.get('지역3') or rec.get('region_3'))
            territory_owner = _str(rec.get('담당자') or rec.get('담당') or rec.get('팀')
                                  or rec.get('territory_owner') or rec.get('Ownership'))
            address = _str(rec.get('주소') or rec.get('소재지주소') or rec.get('address'))
            phone = _str(rec.get('전화번호') or rec.get('전화') or rec.get('phone'))

            # 타겟팅 필드 (숫자 1 또는 True 등)
            is_target_val = rec.get('타겟팅') or rec.get('TARGET') or rec.get('is_target')
            is_target = _bool_ish(is_target_val)

            # 우편번호
            zipcode = _str(rec.get('우편번호'))

            # 제품 리스팅 정보 - 실제 Excel 컬럼명 매핑
            product_listing = {}
            for col in PRODUCT_COLUMNS:
                val = _str(rec.get(col))
                if val:
                    product_listing[col] = val
            # 기존 리스팅/listing 키워드 매핑도 유지
            for key in rec:
                if key not in PRODUCT_COLUMNS and ('리스팅' in str(key) or 'listing' in str(key).lower()):
                    val = _str(rec[key])
                    if val:
                        product_listing[key] = val

            custom_props = {}
            if product_listing:
                custom_props['product_listing'] = product_listing
            if zipcode:
                custom_props['우편번호'] = zipcode

            # 비급여금액, 홈페이지 등 추가 정보
            for extra_key in ['비급여금액', '비급여수가', '홈페이지', '대표자', '개원년도',
                              '최초 실시일', '진료과목']:
                if rec.get(extra_key):
                    custom_props[extra_key] = _str(rec[extra_key])

            if existing:
                # Update
                if hospital_code and not existing.hospital_code:
                    existing.hospital_code = hospital_code
                if hospital_type:
                    existing.hospital_type = hospital_type
                if region_1:
                    existing.region_1 = region_1
                if region_2:
                    existing.region_2 = region_2
                if region_3:
                    existing.region_3 = region_3
                if territory_owner:
                    existing.territory_owner = territory_owner
                if address:
                    existing.address = address
                if phone:
                    existing.phone = phone
                if is_target:
                    existing.is_target = is_target
                if custom_props:
                    merged = existing.custom_properties or {}
                    merged.update(custom_props)
                    existing.custom_properties = merged
                result.updated += 1
            else:
                # Create
                company = Company(
                    name=name,
                    hospital_code=hospital_code,
                    hospital_type=hospital_type,
                    region_1=region_1,
                    region_2=region_2,
                    region_3=region_3,
                    territory_owner=territory_owner,
                    address=address,
                    phone=phone,
                    is_target=is_target,
                    industry="의료",
                    custom_properties=custom_props,
                )
                db.add(company)
                result.created += 1

        except Exception as e:
            result.errors.append(f"행 처리 오류: {str(e)[:100]}")
            result.skipped += 1
            try:
                await db.rollback()
            except Exception:
                pass

    try:
        await db.flush()
    except Exception:
        await db.rollback()
    return SuccessResponse(data=result)


@router.post("/hospitals-territory", response_model=SuccessResponse)
async def import_hospitals_territory(
    file: UploadFile = File(...),
    sheet_name: Optional[str] = Form("병원DB"),
    db: AsyncSession = Depends(get_db),
):
    """병원DB Territory 임포트 (side-by-side 레이아웃: 담당자|병원명|구분 | 담당자|병원명|구분)"""
    try:
        file_bytes = await file.read()
        rows = _parse_excel_raw(file_bytes, sheet_name)
    except Exception as e:
        return ErrorResponse(message=f"파일 파싱 실패: {str(e)}")

    result = ImportResult(total=0)

    # 병원DB는 side-by-side 레이아웃: 담당자, 병원명, 구분 | 빈칸 | 담당자, 병원명, 구분
    # 헤더 행 찾기
    header_row = None
    for i, row in enumerate(rows[:5]):
        vals = [v for v in row if isinstance(v, str)]
        if '담당자' in vals and '병원명' in vals:
            header_row = i
            break

    if header_row is None:
        return ErrorResponse(message="헤더를 찾을 수 없습니다 (담당자, 병원명, 구분)")

    for row in rows[header_row + 1:]:
        # 왼쪽 블록 (col 0~2 또는 1~3)
        pairs = []
        # 행에서 담당자+병원명+구분 쌍 추출
        vals = list(row)
        i = 0
        while i < len(vals) - 1:
            # 담당자-병원명-구분 패턴 또는 병원명-구분 패턴 찾기
            name = _str(vals[i])
            if name and i + 1 < len(vals):
                next_val = _str(vals[i + 1])
                # 구분값 체크 (상급종합, 종합병원, 병원, 의원)
                type_vals = ['상급종합', '종합병원', '병원', '의원', '상급종합병원']
                if next_val in type_vals:
                    # name은 병원명, next_val은 구분 - 담당자는 이전에
                    # 담당자 확인: 이전 값이 Harry/Chloe 등이면 담당자
                    owner = None
                    if i > 0:
                        prev = _str(vals[i - 1])
                        if prev and prev not in type_vals and len(prev) < 15:
                            owner = prev
                    pairs.append((owner, name, next_val))
                    i += 2
                    continue
                elif name in ('Harry', 'Chloe', 'Harry(n)'):
                    # 담당자 → 다음이 병원명
                    hospital_name = next_val
                    h_type = _str(vals[i + 2]) if i + 2 < len(vals) else None
                    if hospital_name:
                        pairs.append((name, hospital_name, h_type))
                    i += 3
                    continue
            i += 1

        for owner, hosp_name, h_type in pairs:
            if not hosp_name:
                continue
            result.total += 1

            try:
                q = await db.execute(
                    select(Company).where(Company.name == hosp_name)
                )
                existing = q.scalar_one_or_none()

                # 담당자 정리 (Harry(n) → Harry)
                if owner:
                    owner = owner.replace('(n)', '').replace('(N)', '').strip()

                if existing:
                    if owner:
                        existing.territory_owner = owner
                    if h_type:
                        existing.hospital_type = h_type
                    result.updated += 1
                else:
                    company = Company(
                        name=hosp_name,
                        hospital_type=h_type,
                        territory_owner=owner,
                        industry="의료",
                    )
                    db.add(company)
                    result.created += 1
            except Exception as e:
                result.errors.append(f"{hosp_name}: {str(e)[:80]}")
                result.skipped += 1

    try:
        await db.flush()
    except Exception:
        await db.rollback()
    return SuccessResponse(data=result)


# ────────────────── Phase 2: HCP 임포트 ──────────────────


@router.post("/hcps", response_model=SuccessResponse)
async def import_hcps(
    file: UploadFile = File(...),
    sheet_name: Optional[str] = Form(None),
    db: AsyncSession = Depends(get_db),
):
    """HCP(의료진) 벌크 임포트 (병원명으로 company 매칭)"""
    try:
        file_bytes = await file.read()
        records = _parse_excel(file_bytes, sheet_name)
    except Exception as e:
        return ErrorResponse(message=f"파일 파싱 실패: {str(e)}")

    result = ImportResult(total=len(records))

    for rec in records:
        try:
            # email unique constraint로 인한 autoflush 에러 방지
            # 각 레코드 처리 전 pending 변경사항 flush
            try:
                await db.flush()
            except Exception:
                await db.rollback()
            name = _str(rec.get('의료진') or rec.get('성명') or rec.get('교수')
                       or rec.get('고객') or rec.get('처방의') or rec.get('의사명')
                       or rec.get('대표자') or rec.get('이름'))
            if not name:
                result.skipped += 1
                continue

            hcp_code = _str(rec.get('고유키값') or rec.get('hcp_code'))
            hospital_name = _str(rec.get('병원명') or rec.get('병원') or rec.get('소속')
                                or rec.get('실시기관명') or rec.get('처방 기관'))

            # 병원 매칭 (부분매칭 포함)
            company_id = None
            if hospital_name:
                # 정확히 매칭
                q = await db.execute(
                    select(Company.id).where(Company.name == hospital_name)
                )
                company_id = q.scalar_one_or_none()
                # 부분매칭 시도
                if not company_id:
                    q = await db.execute(
                        select(Company.id).where(Company.name.contains(hospital_name)).limit(1)
                    )
                    company_id = q.scalar_one_or_none()

            # 기존 HCP 찾기 (hcp_code 우선, 없으면 이름+병원)
            existing = None
            if hcp_code:
                q = await db.execute(
                    select(Contact).where(Contact.hcp_code == hcp_code)
                )
                existing = q.scalar_one_or_none()

            if not existing and company_id:
                q = await db.execute(
                    select(Contact).where(
                        Contact.first_name == name,
                        Contact.company_id == company_id,
                    )
                )
                existing = q.scalar_one_or_none()

            # company_id가 없을 때만 이름으로 fallback (동명이인 다른 병원 방지)
            if not existing and not company_id:
                q = await db.execute(
                    select(Contact).where(Contact.first_name == name)
                )
                existing = q.scalar_one_or_none()

            department = _str(rec.get('진료과') or rec.get('소속임상과') or rec.get('과')
                            or rec.get('처방과') or rec.get('department') or rec.get('Speciality')
                            or rec.get('진료과목'))
            sub_specialty = _str(rec.get('세부전공') or rec.get('sub_specialty'))
            title_position = _str(rec.get('직급') or rec.get('직책') or rec.get('title_position')
                                 or rec.get('고객타입'))
            license_number = _str(rec.get('면허번호') or rec.get('면허번호(취득년도)')
                                 or rec.get('license_number'))
            phone = _str(rec.get('휴대폰') or rec.get('전화') or rec.get('phone')
                        or rec.get('환자 연락처'))
            email = _str(rec.get('E-mail') or rec.get('이메일') or rec.get('email'))
            territory_owner = _str(rec.get('담당자') or rec.get('담당') or rec.get('팀'))
            product = _str(rec.get('제품'))

            # 임상 메트릭 - 실제 Excel 컬럼명으로 매핑 (줄바꿈 제거됨)
            clinical_metrics = {}
            metric_keys = {
                '1주  외래': 'weekly_outpatient',
                '1주 외래': 'weekly_outpatient',
                '1타임  환자(Ave.)': 'patients_per_session',
                '1타임 환자(Ave.)': 'patients_per_session',
                '1타임  초진(Ave.)': 'new_patients_per_session',
                '1타임 초진(Ave.)': 'new_patients_per_session',
                '월  총 환자': 'monthly_total_patients',
                '월 총 환자': 'monthly_total_patients',
                'MCI&AD Poten.': 'mci_ad_potential',
                'MCI&AD Pts.': 'mci_ad_patients',
                'aMCI': 'amci',
                'AD': 'ad',
                '제품A Poten.': 'producta_potential',
                '제품A Rx.(Ave)': 'producta_rx_avg',
                'M/S': 'market_share',
                '월  초진': 'monthly_new_patients',
                '월 초진': 'monthly_new_patients',
                '월  초진 비율': 'new_patient_ratio',
                '월 초진 비율': 'new_patient_ratio',
                '신환대기': 'new_patient_wait',
                'Active User': 'active_user',
                '처방 시작월': 'first_rx_month',
                '처방 건수': 'rx_count',
            }
            for k, eng_k in metric_keys.items():
                val = rec.get(k)
                if val is not None:
                    # datetime → string for JSON serialization
                    from datetime import datetime as _dt
                    if isinstance(val, _dt):
                        val = val.strftime('%Y-%m-%d')
                    clinical_metrics[eng_k] = val

            custom_props = {}
            if clinical_metrics:
                custom_props['clinical_metrics'] = clinical_metrics
            if product:
                custom_props['제품'] = product

            # 추가 정보
            for extra_key in ['고객타입', 'Call 진행여부', '해외학회', 'IIT', '제품설명회',
                              '구분', '비고', '등록현황', '등록완료일자']:
                val = rec.get(extra_key)
                if val is not None:
                    custom_props[extra_key] = _str(val)

            if existing:
                if hcp_code and not existing.hcp_code:
                    existing.hcp_code = hcp_code
                if company_id:
                    existing.company_id = company_id
                if department:
                    existing.department = department
                if sub_specialty:
                    existing.sub_specialty = sub_specialty
                if title_position:
                    existing.title_position = title_position
                if license_number:
                    existing.license_number = license_number
                if phone:
                    existing.phone = phone
                if email and not existing.email:
                    # email unique 보호
                    dup = await db.execute(select(Contact.id).where(Contact.email == email))
                    if not dup.scalar_one_or_none():
                        existing.email = email
                if territory_owner:
                    existing.owner_slack_id = territory_owner
                if custom_props:
                    merged = existing.custom_properties or {}
                    merged.update(custom_props)
                    existing.custom_properties = merged
                result.updated += 1
            else:
                # 새 연락처 생성 시 email unique 보호
                safe_email = email
                if safe_email:
                    dup = await db.execute(select(Contact.id).where(Contact.email == safe_email))
                    if dup.scalar_one_or_none():
                        safe_email = None
                contact = Contact(
                    first_name=name,
                    last_name="",
                    email=safe_email,
                    phone=phone,
                    company_id=company_id,
                    owner_slack_id=territory_owner,
                    hcp_code=hcp_code,
                    department=department,
                    sub_specialty=sub_specialty,
                    title_position=title_position,
                    license_number=license_number,
                    source="excel_import",
                    custom_properties=custom_props,
                )
                db.add(contact)
                result.created += 1

        except Exception as e:
            result.errors.append(f"행 처리 오류: {str(e)[:100]}")
            result.skipped += 1
            try:
                await db.rollback()
            except Exception:
                pass
            try:
                await db.rollback()
            except Exception:
                pass

    try:
        await db.flush()
    except Exception:
        await db.rollback()
    return SuccessResponse(data=result)


@router.post("/doctor-info", response_model=SuccessResponse)
async def import_doctor_info(
    file: UploadFile = File(...),
    sheet_name: Optional[str] = Form(None),
    db: AsyncSession = Depends(get_db),
):
    """의료진정보 머지 (이름+병원 매칭 → 전화/이메일 업데이트)"""
    try:
        file_bytes = await file.read()
        records = _parse_excel(file_bytes, sheet_name)
    except Exception as e:
        return ErrorResponse(message=f"파일 파싱 실패: {str(e)}")

    result = ImportResult(total=len(records))

    for rec in records:
        try:
            name = _str(rec.get('성명') or rec.get('의료진'))
            hospital_name = _str(rec.get('실시기관명') or rec.get('병원명'))
            if not name:
                result.skipped += 1
                continue

            # 이름+병원으로 매칭
            existing = None
            if hospital_name:
                q = await db.execute(
                    select(Contact).join(Company, Contact.company_id == Company.id).where(
                        Contact.first_name == name,
                        Company.name == hospital_name,
                    )
                )
                existing = q.scalar_one_or_none()

            if not existing:
                q = await db.execute(
                    select(Contact).where(Contact.first_name == name)
                )
                existing = q.scalar_one_or_none()

            if existing:
                phone = _str(rec.get('휴대폰') or rec.get('전화'))
                email = _str(rec.get('E-mail') or rec.get('이메일'))
                department = _str(rec.get('소속임상과') or rec.get('진료과'))
                sub_specialty = _str(rec.get('세부전공'))
                title_position = _str(rec.get('직책') or rec.get('직급'))
                license_number = _str(rec.get('면허번호(취득년도)') or rec.get('면허번호'))

                if phone:
                    existing.phone = phone
                if email:
                    existing.email = email
                if department:
                    existing.department = department
                if sub_specialty:
                    existing.sub_specialty = sub_specialty
                if title_position:
                    existing.title_position = title_position
                if license_number:
                    existing.license_number = license_number

                # 등록 현황
                reg_status = _str(rec.get('등록현황'))
                reg_date = _str(rec.get('등록완료일자'))
                note = _str(rec.get('비고'))
                role = _str(rec.get('구분'))  # 실시의사
                if reg_status or role:
                    merged = existing.custom_properties or {}
                    if reg_status:
                        merged['등록현황'] = reg_status
                    if reg_date:
                        merged['등록완료일자'] = reg_date
                    if note:
                        merged['비고'] = note
                    if role:
                        merged['구분'] = role
                    existing.custom_properties = merged

                result.updated += 1
            else:
                # 매칭 안되면 새로 생성
                company_id = None
                if hospital_name:
                    q = await db.execute(
                        select(Company.id).where(Company.name == hospital_name)
                    )
                    company_id = q.scalar_one_or_none()

                phone = _str(rec.get('휴대폰'))
                email = _str(rec.get('E-mail'))
                department = _str(rec.get('소속임상과'))
                title_position = _str(rec.get('직책'))
                license_number = _str(rec.get('면허번호(취득년도)'))

                contact = Contact(
                    first_name=name,
                    last_name="",
                    email=email,
                    phone=phone,
                    company_id=company_id,
                    department=department,
                    title_position=title_position,
                    license_number=license_number,
                    source="doctor_info_import",
                    custom_properties={
                        k: _str(rec.get(k)) for k in ['등록현황', '등록완료일자', '비고', '구분']
                        if rec.get(k)
                    },
                )
                db.add(contact)
                result.created += 1

        except Exception as e:
            result.errors.append(f"행 처리 오류: {str(e)[:100]}")
            result.skipped += 1
            try:
                await db.rollback()
            except Exception:
                pass

    try:
        await db.flush()
    except Exception:
        await db.rollback()
    return SuccessResponse(data=result)


# ────────────────── Phase 3: 처방 임포트 ──────────────────


@router.post("/prescriptions", response_model=SuccessResponse)
async def import_prescriptions(
    file: UploadFile = File(...),
    sheet_name: Optional[str] = Form(None),
    db: AsyncSession = Depends(get_db),
):
    """처방 데이터 벌크 임포트"""
    try:
        file_bytes = await file.read()
        records = _parse_excel(file_bytes, sheet_name)
    except Exception as e:
        return ErrorResponse(message=f"파일 파싱 실패: {str(e)}")

    result = ImportResult(total=len(records))

    for rec in records:
        try:
            prescription_code = _str(rec.get('처방 ID') or rec.get('prescription_code'))
            hospital_name = _str(rec.get('병원') or rec.get('병원명') or rec.get('처방 기관'))
            doctor_name = _str(rec.get('처방의') or rec.get('의사명'))

            if not hospital_name and not doctor_name and not prescription_code:
                result.skipped += 1
                continue

            # 병원 매칭
            hospital_id = None
            if hospital_name:
                # 처방 기관에서 과 제거 (예: "OO병원 정신건강의학과" → "OO병원")
                clean_name = hospital_name.split(' ')[0] if ' ' in hospital_name else hospital_name
                q = await db.execute(
                    select(Company.id).where(Company.name == hospital_name)
                )
                hospital_id = q.scalar_one_or_none()
                if not hospital_id:
                    q = await db.execute(
                        select(Company.id).where(Company.name == clean_name)
                    )
                    hospital_id = q.scalar_one_or_none()
                if not hospital_id:
                    q = await db.execute(
                        select(Company.id).where(Company.name.contains(clean_name)).limit(1)
                    )
                    hospital_id = q.scalar_one_or_none()

            # 의사 매칭
            doctor_id = None
            if doctor_name and isinstance(doctor_name, str):
                query = select(Contact.id).where(Contact.first_name == doctor_name)
                if hospital_id:
                    query = query.where(Contact.company_id == hospital_id)
                q = await db.execute(query)
                doctor_id = q.scalar_one_or_none()
                if not doctor_id:
                    q = await db.execute(
                        select(Contact.id).where(Contact.first_name == doctor_name).limit(1)
                    )
                    doctor_id = q.scalar_one_or_none()

            # 중복 체크
            existing = None
            if prescription_code:
                q = await db.execute(
                    select(Prescription).where(Prescription.prescription_code == prescription_code)
                )
                existing = q.scalar_one_or_none()

            session_number = _int(rec.get('처방 회차') or rec.get('session_number')) or 1
            platform = _str(rec.get('처방 플랫폼') or rec.get('platform'))
            patient_id = _str(rec.get('환자 ID') or rec.get('patient_id'))
            prescription_type = _str(rec.get('처방과') or rec.get('구분')
                                    or rec.get('과') or rec.get('prescription_type'))
            owner = _str(rec.get('담당자'))

            # 날짜 처리
            from datetime import datetime as dt
            prescribed_date = rec.get('처방 코드 생성일') or rec.get('처방 코드 전송 일시') or rec.get('처방일')
            if prescribed_date and not isinstance(prescribed_date, dt):
                prescribed_date = None

            activated_date = rec.get('처방 코드 활성화일') or rec.get('처방 코드 활성화 일시') or rec.get('활성화일')
            if activated_date and not isinstance(activated_date, dt):
                activated_date = None

            custom_props = {}
            if owner:
                custom_props['담당자'] = owner
            # 년도/월/일
            year = _int(rec.get('년도'))
            month = _int(rec.get('월'))
            if year:
                custom_props['year'] = year
            if month:
                custom_props['month'] = month

            if existing:
                if hospital_id:
                    existing.hospital_id = hospital_id
                if doctor_id:
                    existing.doctor_id = doctor_id
                if session_number:
                    existing.session_number = session_number
                if activated_date:
                    existing.activated_date = activated_date
                if prescription_type and not existing.prescription_type:
                    existing.prescription_type = prescription_type
                result.updated += 1
            else:
                rx = Prescription(
                    prescription_code=prescription_code,
                    session_number=session_number,
                    platform=platform,
                    hospital_id=hospital_id,
                    doctor_id=doctor_id,
                    patient_id=patient_id,
                    prescription_type=prescription_type,
                    prescribed_date=prescribed_date,
                    activated_date=activated_date,
                    custom_properties=custom_props,
                )
                db.add(rx)
                result.created += 1

        except Exception as e:
            result.errors.append(f"행 처리 오류: {str(e)[:100]}")
            result.skipped += 1
            try:
                await db.rollback()
            except Exception:
                pass

    try:
        await db.flush()
    except Exception:
        await db.rollback()
    return SuccessResponse(data=result)


@router.post("/compliance", response_model=SuccessResponse)
async def import_compliance(
    file: UploadFile = File(...),
    sheet_name: Optional[str] = Form(None),
    db: AsyncSession = Depends(get_db),
):
    """환자 순응도 데이터 임포트 (사용자 일차별 훈련 참여도 / 재처방 리스트)"""
    try:
        file_bytes = await file.read()
        records = _parse_excel(file_bytes, sheet_name)
    except Exception as e:
        return ErrorResponse(message=f"파일 파싱 실패: {str(e)}")

    result = ImportResult(total=len(records))

    for rec in records:
        try:
            patient_id = _str(rec.get('환자 ID') or rec.get('patient_id'))
            if not patient_id:
                result.skipped += 1
                continue

            # 이미 존재하는지 확인
            q = await db.execute(
                select(PatientCompliance).where(PatientCompliance.patient_id == patient_id)
            )
            existing = q.scalar_one_or_none()

            # 사용일수 = total_sessions
            total_sessions = _int(rec.get('사용일수')) or 0

            # 일차별 참여도에서 completed_sessions 계산 (1.0 = 완료)
            completed = 0
            for key in rec:
                if '일차' in str(key):
                    val = _float(rec[key])
                    if val is not None and val >= 0.5:
                        completed += 1

            # 순응도
            recent_avg = _float(rec.get('최근 7일 평균참여도'))
            compliance_rate = recent_avg if recent_avg is not None else (
                completed / total_sessions if total_sessions > 0 else 0.0
            )

            # 재처방 리스트 형식의 순응도
            compliance_1 = _float(rec.get('1회 순응도'))
            compliance_2 = _float(rec.get('2회 순응도'))

            # 병원 매칭
            hospital_name = _str(rec.get('처방 기관'))
            hospital_id = None
            if hospital_name:
                clean_name = hospital_name.split(' ')[0] if ' ' in hospital_name else hospital_name
                q = await db.execute(
                    select(Company.id).where(Company.name.contains(clean_name)).limit(1)
                )
                hospital_id = q.scalar_one_or_none()

            # 의사 매칭
            doctor_name = _str(rec.get('처방의'))
            doctor_id = None
            if doctor_name:
                q = await db.execute(
                    select(Contact.id).where(Contact.first_name == doctor_name).limit(1)
                )
                doctor_id = q.scalar_one_or_none()

            custom = {}
            if compliance_1 is not None:
                custom['1회_순응도'] = compliance_1
            if compliance_2 is not None:
                custom['2회_순응도'] = compliance_2
            change = _float(rec.get('변화률'))
            if change is not None:
                custom['변화률'] = change
            # 환자 정보 (재처방 리스트)
            for extra in ['이름', '성별', '생년월일', '환자 연락처']:
                val = _str(rec.get(extra))
                if val:
                    custom[extra] = val

            if existing:
                existing.total_sessions = max(existing.total_sessions, total_sessions)
                existing.completed_sessions = max(existing.completed_sessions, completed)
                existing.compliance_rate = compliance_rate
                if hospital_id:
                    existing.hospital_id = hospital_id
                if doctor_id:
                    existing.doctor_id = doctor_id
                if custom:
                    merged = existing.custom_properties or {}
                    merged.update(custom)
                    existing.custom_properties = merged
                result.updated += 1
            else:
                comp = PatientCompliance(
                    patient_id=patient_id,
                    hospital_id=hospital_id,
                    doctor_id=doctor_id,
                    total_sessions=total_sessions,
                    completed_sessions=completed,
                    compliance_rate=compliance_rate,
                    custom_properties=custom,
                )
                db.add(comp)
                result.created += 1

        except Exception as e:
            result.errors.append(f"행 처리 오류: {str(e)[:100]}")
            result.skipped += 1
            try:
                await db.rollback()
            except Exception:
                pass

    try:
        await db.flush()
    except Exception:
        await db.rollback()
    return SuccessResponse(data=result)


# ────────────────── Phase 4: 매출 임포트 ──────────────────


@router.post("/sales", response_model=SuccessResponse)
async def import_sales(
    file: UploadFile = File(...),
    sheet_name: Optional[str] = Form(None),
    db: AsyncSession = Depends(get_db),
):
    """매출 데이터 벌크 임포트"""
    try:
        file_bytes = await file.read()
        # Sales Tracking은 col 0~15만 사용 (col 20+ 는 pivot 테이블)
        records = _parse_excel(file_bytes, sheet_name, max_col=16)
    except Exception as e:
        return ErrorResponse(message=f"파일 파싱 실패: {str(e)}")

    result = ImportResult(total=len(records))

    for rec in records:
        try:
            year = _int(rec.get('연도') or rec.get('year'))
            month = _int(rec.get('월') or rec.get('month'))
            if not year or not month:
                result.skipped += 1
                continue

            company_name = _str(rec.get('회사') or rec.get('병원') or rec.get('병원명'))
            company_id = None
            if company_name:
                q = await db.execute(
                    select(Company.id).where(Company.name == company_name)
                )
                company_id = q.scalar_one_or_none()

            product = _str(rec.get('제품') or rec.get('product'))
            channel = _str(rec.get('채널(거래처)') or rec.get('채널') or rec.get('channel'))
            quantity = _int(rec.get('수량') or rec.get('quantity')) or 0
            # 공급가가 매출 대신 사용되는 경우
            unit_price = _float(rec.get('공급가') or rec.get('unit_price')) or 0.0
            revenue = _float(rec.get('매출(-VAT)') or rec.get('매출') or rec.get('revenue'))
            # 매출이 없으면 공급가를 매출로 사용
            if revenue is None or revenue == 0.0:
                revenue = unit_price
            revenue_recognized = _bool_ish(rec.get('매출인식') or rec.get('revenue_recognized'))
            payment_received = _bool_ish(rec.get('입금여부') or rec.get('payment_received'))
            ownership = _str(rec.get('Ownership') or rec.get('ownership'))

            from datetime import datetime as dt
            payment_date = rec.get('입금일')
            if payment_date and not isinstance(payment_date, dt):
                payment_date = None

            custom_props = {}
            for extra_key in ['품의여부 (매출)', '품의여부(매출)', '세금계산서 발행',
                              '세금계산서발행', '기타']:
                val = rec.get(extra_key)
                if val is not None:
                    custom_props[extra_key.strip()] = _str(val)

            sale = SalesTransaction(
                year=year,
                month=month,
                company_id=company_id,
                product=product,
                channel=channel,
                quantity=quantity,
                unit_price=unit_price,
                revenue=revenue,
                revenue_recognized=revenue_recognized,
                payment_received=payment_received,
                payment_date=payment_date,
                ownership=ownership,
                custom_properties=custom_props,
            )
            db.add(sale)
            result.created += 1

        except Exception as e:
            result.errors.append(f"행 처리 오류: {str(e)[:100]}")
            result.skipped += 1
            try:
                await db.rollback()
            except Exception:
                pass

    try:
        await db.flush()
    except Exception:
        await db.rollback()
    return SuccessResponse(data=result)


# ────────────────── Phase 5: 제품 리스팅/KOL/계약 임포트 ──────────────────


@router.post("/product-listings", response_model=SuccessResponse)
async def import_product_listings(
    file: UploadFile = File(...),
    sheet_name: Optional[str] = Form(None),
    db: AsyncSession = Depends(get_db),
):
    """제품 리스팅 벌크 임포트"""
    try:
        file_bytes = await file.read()
        records = _parse_excel(file_bytes, sheet_name)
    except Exception as e:
        return ErrorResponse(message=f"파일 파싱 실패: {str(e)}")

    result = ImportResult(total=len(records))

    for rec in records:
        try:
            hospital_name = _str(rec.get('병원명') or rec.get('병원') or rec.get('요양기관명'))
            product = _str(rec.get('제품') or rec.get('product'))
            if not hospital_name:
                result.skipped += 1
                continue

            # 병원 매칭 (부분매칭)
            q = await db.execute(
                select(Company.id).where(Company.name == hospital_name)
            )
            company_id = q.scalar_one_or_none()
            if not company_id:
                q = await db.execute(
                    select(Company.id).where(Company.name.contains(hospital_name)).limit(1)
                )
                company_id = q.scalar_one_or_none()
            if not company_id:
                # 병원 자동 생성
                new_company = Company(name=hospital_name, industry="의료")
                db.add(new_company)
                await db.flush()
                company_id = new_company.id

            status = _str(rec.get('상태') or rec.get('리스팅 상태') or rec.get('Listing Status')
                         or rec.get('On-going') or rec.get('status'))
            # 상태값 정규화
            if status:
                status_map = {
                    '완료': 'done', '진행중': 'in_progress', '계획': 'pending',
                    '구두컨펌': 'verbal_confirm', 'done': 'done',
                    'listing': 'done', 'on-going': 'in_progress',
                }
                status = status_map.get(status.lower(), status) if isinstance(status, str) else status
            else:
                status = "pending"

            pipeline_stage = _str(rec.get('단계') or rec.get('pipeline_stage')
                                 or rec.get('구분'))
            notes = _str(rec.get('비고') or rec.get('notes') or rec.get('기타'))

            # 제품 리스팅용 KOL 정보
            kol_name = _str(rec.get('KoL') or rec.get('KOLs'))

            custom = {}
            if kol_name:
                custom['KOL'] = kol_name
            # Pts 정보
            pts = _str(rec.get('Pts') or rec.get('Pts. 6주'))
            if pts:
                custom['Pts'] = pts

            listing = ProductListing(
                company_id=company_id,
                product=product or "제품A",
                status=status,
                pipeline_stage=pipeline_stage,
                notes=notes,
                custom_properties=custom,
            )
            db.add(listing)
            result.created += 1

        except Exception as e:
            result.errors.append(f"행 처리 오류: {str(e)[:100]}")
            result.skipped += 1
            try:
                await db.rollback()
            except Exception:
                pass

    try:
        await db.flush()
    except Exception:
        await db.rollback()
    return SuccessResponse(data=result)


@router.post("/kol-plans", response_model=SuccessResponse)
async def import_kol_plans(
    file: UploadFile = File(...),
    sheet_name: Optional[str] = Form(None),
    db: AsyncSession = Depends(get_db),
):
    """KOL 계획 벌크 임포트 (DTx KOLs Plan: 병원당 여러 의사 — PM, RM, NR, NP)"""
    try:
        file_bytes = await file.read()
        records = _parse_excel(file_bytes, sheet_name)
    except Exception as e:
        return ErrorResponse(message=f"파일 파싱 실패: {str(e)}")

    result = ImportResult(total=len(records))

    # DTx KOLs Plan 형식: 번호, 병원, 담당자, PM, 외래타임, RM, 외래타임2, NR*, 외래타임3, NP*, 외래타임4
    # 또는 ILD/COPD Target 형식: 연구회, 소속, 지역, 교수, 담당자, Call진행여부

    for rec in records:
        try:
            hospital_name = _str(rec.get('병원') or rec.get('소속') or rec.get('병원명'))
            territory_owner = _str(rec.get('담당자') or rec.get('담당'))

            # 병원 매칭
            company_id = None
            if hospital_name:
                q = await db.execute(
                    select(Company.id).where(Company.name == hospital_name)
                )
                company_id = q.scalar_one_or_none()
                if not company_id:
                    q = await db.execute(
                        select(Company.id).where(Company.name.contains(hospital_name)).limit(1)
                    )
                    company_id = q.scalar_one_or_none()

            # DTx KOLs Plan 형식 - 여러 의사 (PM, RM, NR*, NP*)
            doctor_roles = ['PM', 'RM', 'NR*', 'NP*', 'NR', 'NP']
            schedule_keys = ['외래타임', '외래타임 2', '외래타임2', '외래타임 3', '외래타임3',
                            '외래타임 4', '외래타임4']
            found_doctors = False

            for role in doctor_roles:
                doctor_name = _str(rec.get(role))
                if not doctor_name:
                    continue
                found_doctors = True

                # 쉼표로 구분된 복수 의사 처리 (예: "양동원, 윤보라")
                names = [n.strip() for n in doctor_name.split(',') if n.strip()]

                for dname in names:
                    doctor_id = None
                    q = await db.execute(
                        select(Contact.id).where(Contact.first_name == dname).limit(1)
                    )
                    doctor_id = q.scalar_one_or_none()

                    # 외래 스케줄 찾기
                    schedule = {}
                    role_idx = doctor_roles.index(role) if role in doctor_roles else 0
                    if role_idx < len(schedule_keys):
                        sched_val = _str(rec.get(schedule_keys[role_idx]))
                        if sched_val:
                            schedule['외래'] = sched_val

                    plan_type = _str(rec.get('연구회') or rec.get('plan_type'))
                    engagement = _str(rec.get('Call 진행여부') or rec.get('engagement_status'))

                    kol = KOLPlan(
                        company_id=company_id,
                        doctor_id=doctor_id,
                        plan_type=plan_type or role.replace('*', ''),
                        target_product=_str(rec.get('제품')) or "제품A",
                        clinic_schedule=schedule,
                        engagement_status=engagement or "planned",
                        notes=f"{hospital_name} - {dname} ({role})" if hospital_name else dname,
                        custom_properties={'담당자': territory_owner, '역할': role} if territory_owner else {'역할': role},
                    )
                    db.add(kol)
                    result.created += 1

            # ILD/COPD Target 형식 (단일 교수)
            if not found_doctors:
                doctor_name = _str(rec.get('교수') or rec.get('의료진') or rec.get('Key Doctor'))
                if not doctor_name:
                    result.skipped += 1
                    continue

                doctor_id = None
                q = await db.execute(
                    select(Contact.id).where(Contact.first_name == doctor_name).limit(1)
                )
                doctor_id = q.scalar_one_or_none()

                plan_type = _str(rec.get('연구회') or rec.get('plan_type'))
                engagement = _str(rec.get('Call 진행여부'))

                kol = KOLPlan(
                    company_id=company_id,
                    doctor_id=doctor_id,
                    plan_type=plan_type,
                    target_product=_str(rec.get('제품')),
                    engagement_status=engagement or "planned",
                    custom_properties={'담당자': territory_owner, '지역': _str(rec.get('지역'))} if territory_owner else {},
                )
                db.add(kol)
                result.created += 1

        except Exception as e:
            result.errors.append(f"행 처리 오류: {str(e)[:100]}")
            result.skipped += 1
            try:
                await db.rollback()
            except Exception:
                pass

    try:
        await db.flush()
    except Exception:
        await db.rollback()
    return SuccessResponse(data=result)


@router.post("/hospital-contracts", response_model=SuccessResponse)
async def import_hospital_contracts(
    file: UploadFile = File(...),
    sheet_name: Optional[str] = Form(None),
    db: AsyncSession = Depends(get_db),
):
    """병원 계약 벌크 임포트 (유통A 병원계약현황 포함)"""
    try:
        file_bytes = await file.read()
        records = _parse_excel(file_bytes, sheet_name)
    except Exception as e:
        return ErrorResponse(message=f"파일 파싱 실패: {str(e)}")

    result = ImportResult(total=len(records))

    for rec in records:
        try:
            hospital_name = _str(rec.get('병원명') or rec.get('병원') or rec.get('병원 LIST'))
            if not hospital_name:
                result.skipped += 1
                continue

            q = await db.execute(
                select(Company.id).where(Company.name == hospital_name)
            )
            company_id = q.scalar_one_or_none()
            if not company_id:
                q = await db.execute(
                    select(Company.id).where(Company.name.contains(hospital_name)).limit(1)
                )
                company_id = q.scalar_one_or_none()
            if not company_id:
                # 자동 생성
                new_company = Company(name=hospital_name, industry="의료")
                db.add(new_company)
                await db.flush()
                company_id = new_company.id

            product = _str(rec.get('제품') or rec.get('product')) or "유통A"
            contract_status = _str(rec.get('계약상태') or rec.get('contract_status')
                                  or rec.get('구매계약'))

            # 유통A 파이프라인 단계 → custom_properties
            pipeline = {}
            pipeline_cols = [
                'NECA 등록여부', 'NECA등록여부', '보험수가신설',
                '처방코드신설', 'EMR연동', '구매계약',
            ]
            for col in pipeline_cols:
                val = _str(rec.get(col))
                if val:
                    pipeline[col] = val

            # 환자 정보
            pts_6 = _str(rec.get('Pts. 6주') or rec.get('Pts.6주'))
            pts_12 = _str(rec.get('Pts. 12주') or rec.get('Pts.12주'))

            custom = {}
            if pipeline:
                custom['pipeline_steps'] = pipeline
            if pts_6:
                custom['Pts_6주'] = pts_6
            if pts_12:
                custom['Pts_12주'] = pts_12

            # 담당자
            owner = _str(rec.get('담당') or rec.get('담당자'))
            if owner:
                custom['담당자'] = owner

            # Specialty / KOLs
            specialty = _str(rec.get('Speciality'))
            kols = _str(rec.get('KOLs'))
            if specialty:
                custom['Speciality'] = specialty
            if kols:
                custom['KOLs'] = kols

            # 전체 상태 결정
            if not contract_status:
                # pipeline 단계에서 가장 진행된 상태로
                if pipeline:
                    all_vals = list(pipeline.values())
                    if any('완료' in v for v in all_vals):
                        contract_status = "in_progress"
                    elif any('진행' in v for v in all_vals):
                        contract_status = "in_progress"
                    else:
                        contract_status = "pending"
                else:
                    contract_status = "pending"

            contract_value = _float(rec.get('계약금액') or rec.get('contract_value')) or 0.0
            notes = _str(rec.get('비고') or rec.get('notes') or rec.get('기타'))

            contract = HospitalContract(
                company_id=company_id,
                product=product,
                contract_status=contract_status,
                contract_value=contract_value,
                notes=notes,
                custom_properties=custom,
            )
            db.add(contract)
            result.created += 1

        except Exception as e:
            result.errors.append(f"행 처리 오류: {str(e)[:100]}")
            result.skipped += 1
            try:
                await db.rollback()
            except Exception:
                pass

    try:
        await db.flush()
    except Exception:
        await db.rollback()
    return SuccessResponse(data=result)
