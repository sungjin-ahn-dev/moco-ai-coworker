"""AICC 시나리오 xlsx 로더.

시나리오 xlsx에서 G-code별 TTS 멘트 dict와 FAQ Q&A 텍스트를 추출한다.
런타임 고정 경로를 먼저 찾고, 없으면 프로젝트 루트의 파일로 폴백한다.
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Optional

logger = logging.getLogger(__name__)

_RUNTIME_PATH = Path("/home/user/MOCO_DATA/aicc_scenario.xlsx")
_PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent.parent
_FALLBACK_PATH = _PROJECT_ROOT / "AICC_세팅_시나리오_v1.2.xlsx"

_cache: dict = {"path": None, "g_codes": None, "faq_text": None, "faq_items": None}


def find_xlsx() -> Optional[Path]:
    for p in (_RUNTIME_PATH, _FALLBACK_PATH):
        if p.exists():
            return p
    return None


def _parse(path: Path) -> tuple[dict[str, str], str, list[dict]]:
    """xlsx에서 G-code 멘트 + FAQ 추출."""
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True, read_only=True)

    # 01_기본흐름_스크립트
    g_codes: dict[str, str] = {}
    if "01_기본흐름_스크립트" in wb.sheetnames:
        sheet = wb["01_기본흐름_스크립트"]
        rows = list(sheet.iter_rows(values_only=True))
        # 첫 행은 헤더: 단계코드 | 구분 | 상황·트리거 | 봇 발화 스크립트(TTS) | ...
        for row in rows[1:]:
            if not row or not row[0]:
                continue
            code = str(row[0]).strip()
            mentss = str(row[3] or "").strip()
            if code.startswith("G") and mentss:
                g_codes[code] = mentss

    # 02_FAQ_스크립트
    faq_items: list[dict] = []
    if "02_FAQ_스크립트" in wb.sheetnames:
        sheet = wb["02_FAQ_스크립트"]
        rows = list(sheet.iter_rows(values_only=True))
        # 헤더: No. | 카테고리 | 고객 질문 | 봇 답변 스크립트 | 답변 후 처리
        for row in rows[1:]:
            if not row or row[0] is None:
                continue
            try:
                no = int(float(row[0]))
            except (ValueError, TypeError):
                continue
            category = str(row[1] or "").strip()
            question = str(row[2] or "").strip()
            answer = str(row[3] or "").strip()
            if question and answer:
                faq_items.append({"no": no, "category": category, "q": question, "a": answer})

    # FAQ를 system_prompt에 주입할 텍스트로 직렬화
    if faq_items:
        lines = []
        current_cat = None
        for it in faq_items:
            if it["category"] != current_cat:
                current_cat = it["category"]
                lines.append(f"\n[{current_cat}]")
            lines.append(f"Q{it['no']}. {it['q']}")
            lines.append(f"A. {it['a']}")
        faq_text = "\n".join(lines).strip()
    else:
        faq_text = ""

    wb.close()
    return g_codes, faq_text, faq_items


def load(force_reload: bool = False) -> dict:
    """캐시된 시나리오 데이터 반환. 파일 없으면 빈 dict."""
    global _cache
    if not force_reload and _cache["g_codes"] is not None:
        return _cache

    path = find_xlsx()
    if path is None:
        logger.warning(f"[SCENARIO_LOADER] xlsx not found in {_RUNTIME_PATH} or {_FALLBACK_PATH}")
        _cache = {"path": None, "g_codes": {}, "faq_text": "", "faq_items": []}
        return _cache

    try:
        g_codes, faq_text, faq_items = _parse(path)
        _cache = {"path": str(path), "g_codes": g_codes, "faq_text": faq_text, "faq_items": faq_items}
        logger.info(
            f"[SCENARIO_LOADER] loaded from {path}: "
            f"{len(g_codes)} G-codes, {len(faq_items)} FAQ items"
        )
    except Exception as e:
        logger.error(f"[SCENARIO_LOADER] parse failed for {path}: {e}")
        _cache = {"path": str(path), "g_codes": {}, "faq_text": "", "faq_items": []}

    return _cache


def get_g_codes() -> dict[str, str]:
    return load()["g_codes"]


def get_faq_text() -> str:
    return load()["faq_text"]


def get_faq_items() -> list[dict]:
    return load()["faq_items"]
